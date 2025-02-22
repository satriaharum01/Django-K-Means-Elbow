import os
import base64
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import asyncio
from asgiref.sync import sync_to_async
from django.template.loader import render_to_string
from sklearn.cluster import KMeans
from django.template import loader
from django.http import HttpResponse
from django.shortcuts import render
from django.conf import settings
from openpyxl import load_workbook
from django.http import JsonResponse
from sklearn.datasets import make_blobs
from django.conf import settings
from xhtml2pdf import pisa
from io import BytesIO
from django.core.files.base import ContentFile


# Decorators
from ..decorators import admin_required, guru_bk_required

# Form
from .forms import ExcelUploadForm, CenteroidForm

# Models Import
from django.db.models import Max
from django.contrib.auth.models import User
from ..models import m_data

# Set the path to save the plot inside the static folder
elbow_dir = os.path.join(settings.BASE_DIR, "static", "elbow")
kmeans_dir = os.path.join(settings.BASE_DIR, "static", "kmeans")
pdf_dir = os.path.join(settings.BASE_DIR, "static", "pdf")

# Create your views here.
# Get the maximum ID
max_id = m_data.objects.aggregate(Max('id'))['id__max'] or 0

@admin_required
def index(request):
    form = ExcelUploadForm()
    CentForm = CenteroidForm()
    context = {
        "title": "K-Means & Elbow",
        "segment": "Clustering",
        "form": form,
        "centeroidForm": CentForm,
    }

    html_template = loader.get_template("page/peramalan.html")
    return HttpResponse(html_template.render(context, request))


def upload_excel(request):

    if request.method == "POST":
        file = request.FILES["file"]

        # Baca file Excel
        wb = load_workbook(file)
        sheet = wb.active  # Ambil sheet aktif
        data = []

        # Iterasi baris dan kolom
        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        # Kembalikan data JSON
        return JsonResponse({"data": data})


# Inisialisasi centroid secara random
def initialize_centroids_random(X, n_clusters):
    random_indices = np.random.choice(range(X.shape[0]), size=n_clusters, replace=False)
    centroids = X[random_indices]
    return centroids


# Inisialisasi centroid berdasarkan average


def initialize_centroids_average(X, n_clusters):
    # Tentukan jumlah data dan tentukan berapa banyak data per cluster
    num_data = X.shape[0]
    data_per_cluster = num_data // n_clusters

    # Inisialisasi centroid berdasarkan pembagian data
    centroids = []
    for i in range(n_clusters):
        # Tentukan indeks mulai dan indeks akhir untuk setiap cluster
        start_index = i * data_per_cluster
        end_index = (i + 1) * data_per_cluster if i != n_clusters - 1 else num_data

        # Ambil subset data untuk cluster ini
        cluster_data = X[start_index:end_index]

        # Hitung rata-rata (centroid) untuk subset data ini
        centroid = np.mean(cluster_data, axis=0)
        centroids.append(centroid)

    # Mengonversi list centroid ke array numpy
    centroids = np.array(centroids)
    return centroids


def initialize_centroids_average_2(X, n_clusters):
    avg = np.mean(X, axis=0)  # Rata-rata tiap kolom
    perturbations = np.linspace(-1, 1, n_clusters)  # Variasi untuk memisahkan centroid
    centroids = np.array([avg + pert * avg for pert in perturbations])
    return centroids


def euclidean_distance(point1, point2):
    return np.sqrt(np.sum((point1 - point2) ** 2))


def kmeans_step_by_step(request):

    context = {"form": ExcelUploadForm()}

    if request.method == "POST":
        try:
            file = request.FILES["file"]
        except Exception as e:
            context["error"] = "File Excel tidak valid atau kosong."
            return render(request, "page/kmeans_step.html", context)

        file = request.FILES["file"]

        # Membuat DataFrame
        df = pd.read_excel(file)
        if df is None:
            context["error"] = "File Excel tidak valid atau kosong."
            return render(request, "page/kmeans_step.html", context)
        # Pastikan dataset memiliki kolom yang sesuai
        required_columns = ["No", "X", "Y"]

        if not all(col in df.columns for col in required_columns):
            context["error"] = (
                f"Dataset harus memiliki kolom: {', '.join(required_columns)}"
            )
            return render(request, "page/kmeans_step.html", context)

        # Inisialisasi variabel untuk menyimpan langkah
        iteration_steps = []

        # Langkah 1: Inisialisasi centroid
        # Pilih kolom untuk clustering
        X = df[["X", "Y"]].to_numpy()

        try:
            k = int(request.POST.get("clusters", 10))  # Jumlah kluster
            n_iter = int(request.POST.get("iterations", 10))  # Jumlah Iterasi
            init_method = request.POST.get("init_method")  # Metode Insiasi
        except Exception as e:
            context["error"] = "Silahkan Isi Form Inisiasi Cluster"
            return render(request, "page/kmeans_step.html", context)
        # Manual K-Means untuk mencatat hasil per iterasi
        k = int(request.POST.get("clusters", 10))  # Jumlah kluster
        n_iter = int(request.POST.get("iterations", 10))  # Jumlah Iterasi
        init_method = request.POST.get("init_method")  # Metode Insiasi

        if init_method == "random":
            centroids = initialize_centroids_random(X, k)
        elif init_method == "average":
            centroids = initialize_centroids_average(X, k)
        else:
            context["error"] = "init_method must be 'random' or 'average'"
            return render(request, "page/kmeans_step.html", context)

        uploadedData = df.to_dict("records")
        for iteration in range(n_iter):
            # Hitung jarak Euclidean
            distances = np.round(
                np.linalg.norm(X[:, None] - centroids[None, :, :], axis=2), 4
            )
            clusters = np.argmin(distances, axis=1)

            # Catat langkah per iterasi
            combined_data = []
            for data, cl, dst in zip(uploadedData, clusters, distances):
                combined_entry = data.copy()
                combined_entry["clusters"] = cl + 1
                combined_entry["euclidean_distances"] = dst
                combined_data.append(combined_entry)

            step_data = {
                "iteration": iteration + 1,
                "count": k,
                "centroids": np.round(centroids, 2).tolist(),
                "clusters": (clusters + 1).tolist(),
                "euclidean_distances": distances.tolist(),
                "combined_data": combined_data,
            }
            iteration_steps.append(step_data)

            # Update centroid
            new_centroids = np.array(
                [
                    (
                        X[clusters == i].mean(axis=0)
                        if np.any(clusters == i)
                        else centroids[i]
                    )
                    for i in range(k)
                ]
            )
            if np.all(
                new_centroids == centroids
            ):  # Jika tidak ada perubahan centroid, berhenti

                break
            centroids = new_centroids

        # Tambahkan hasil akhir ke DataFrame
        df["Kluster"] = clusters
        df["Euclidean Distance"] = [distances[i, clusters[i]] for i in range(len(X))]

        # Menampilkan hasil ke console
        # print("\n=== Hasil K-Means Clustering ===")
        # for step in iteration_steps:
        #    print(f"\n--- Iterasi {step['iteration']} ---")
        #    print("Centroids:", step['centroids'])
        #    print("Clusters:", step['clusters'])
        #    print("Euclidean Distances:", step['euclidean_distances'])
        #    print("Euclidean Distances:", step['combined_data'])
        # Kirim data langkah-langkah ke template
        return render(request, "page/kmeans_step.html", {"steps": iteration_steps})


def kmeans_step_by_step_2(request):

    context = {"form": ExcelUploadForm()}

    if request.method == "POST":
        file = request.FILES["file"]

        # Membuat DataFrame
        df = pd.read_excel(file)
        # Pastikan dataset memiliki kolom yang sesuai
        required_columns = ["Data 1", "Data 2", "Data 3"]

        if not all(col in df.columns for col in required_columns):
            context["error"] = (
                f"Dataset harus memiliki kolom: {', '.join(required_columns)}"
            )
            return render(request, "page/kmeans_step.html", context)

        # Inisialisasi variabel untuk menyimpan langkah
        steps = []

        # Langkah 1: Inisialisasi centroid
        n_clusters = 3
        max_iter = 3
        X = df[["Data 1", "Data 2", "Data 3"]]
        random_state = 0
        kmeans = KMeans(
            n_clusters=n_clusters,
            random_state=random_state,
            init="random",
            n_init=1,
            max_iter=max_iter,
        )
        kmeans.fit(X)
        centroids = kmeans.cluster_centers_

        steps.append(
            {
                "step": "Inisialisasi centroid",
                "tags": "first",
                "count": n_clusters,
                "details": centroids.tolist(),
                "data": df.to_dict("records"),
                "distance": "no",
            }
        )

        # Iterasi clustering manual
        for i in range(max_iter):  # Maksimal 3 iterasi untuk demo

            # Hitung Euclidean Distance dari setiap siswa ke centroid
            distances_to_centroids = []
            for j in range(len(X)):
                distance = np.array(
                    [euclidean_distance(X.iloc[j], centroid) for centroid in centroids]
                )
                distances_to_centroids.append(distance)

            df["Euclidean_Distances"] = distances_to_centroids

            steps.append(
                {
                    "step": f"Iterasi {i+1} - Hitung Euclidean Distances",
                    "count": n_clusters,
                    "data": df.to_dict("records"),
                    "distance": "yes",
                }
            )
            # Langkah 2: Assign cluster
            labels = kmeans.predict(X)
            df["Kluster"] = labels
            df["Kluster"] += 1

            steps.append(
                {
                    "step": f"Iterasi {i+1} - Hasil cluster",
                    "details": df[["No", "Kluster"]].to_dict("records"),
                    "data": df.to_dict("records"),
                    "distance": "no",
                }
            )

            # Langkah 3: Update centroid
            centroids = (
                df.groupby("Kluster").mean()[["Data 1", "Data 2", "Data 3"]].values
            )

            steps.append(
                {
                    "step": f"Iterasi {i+1} - Update centroid",
                    "tags": "first",
                    "count": n_clusters,
                    "details": centroids.tolist(),
                    "data": df.to_dict("records"),
                    "distance": "no",
                }
            )
            # Update model dengan centroid baru
            kmeans.cluster_centers_ = centroids
            
        # Kirim data langkah-langkah ke template
        return render(request, "page/kmeans_step.html", {"steps": steps})


async def clustering_view(request):
    context = {"form": ExcelUploadForm()}

    if request.method == "POST":
        try:
            file = request.FILES["file"]
        except Exception as e:
            context["error"] = "File Excel tidak valid atau kosong."
            return render(request, "page/kmeans_result.html", context)

        file = request.FILES["file"]

        # Membuat DataFrame
        df = pd.read_excel(file)
        if df is None:
            context["error"] = "File Excel tidak valid atau kosong."
            return render(request, "page/kmeans_result.html", context)
        # Pastikan dataset memiliki kolom yang sesuai
        required_columns = ["No", "X", "Y"]

        if not all(col in df.columns for col in required_columns):
            context["error"] = (
                f"Dataset harus memiliki kolom: {', '.join(required_columns)}"
            )
            return render(request, "page/kmeans_result", context)

        # Inisialisasi variabel untuk menyimpan langkah
        iteration_steps = []

        # Langkah 1: Inisialisasi centroid
        # Pilih kolom untuk clustering
        X = df[["X", "Y"]].to_numpy()

        try:
            k = int(request.POST.get("clusters", 10))  # Jumlah kluster
            n_iter = int(request.POST.get("iterations", 10))  # Jumlah Iterasi
            init_method = request.POST.get("init_method")  # Metode Insiasi
        except Exception as e:
            context["error"] = "Silahkan Isi Form Inisiasi Cluster"
            return render(request, "page/kmeans_result.html", context)
        if k > 10:
            context["error"] = "Maximum Cluster 10"
            return render(request, "page/kmeans_result.html", context)
        # Manual K-Means untuk mencatat hasil per iterasi
        k = int(request.POST.get("clusters", 10))  # Jumlah kluster
        n_iter = int(request.POST.get("iterations", 10))  # Jumlah Iterasi
        init_method = request.POST.get("init_method")  # Metode Insiasi

        if init_method == "random":
            centroids = initialize_centroids_random(X, k)
        elif init_method == "average":
            centroids = initialize_centroids_average(X, k)
        else:
            context["error"] = "init_method must be 'random' or 'average'"
            return render(request, "page/kmeans_result.html", context)

        uploadedData = df.to_dict("records")

        # Inisialisasi variabel untuk menyimpan hasil
        result = []

        for iteration in range(n_iter):
            # Hitung jarak Euclidean
            distances = np.round(
                np.linalg.norm(X[:, None] - centroids[None, :, :], axis=2), 4
            )
            clusters = np.argmin(distances, axis=1)

            # Catat langkah per iterasi
            combined_data = []
            for data, cl, dst in zip(uploadedData, clusters, distances):
                combined_entry = data.copy()
                combined_entry["clusters"] = cl + 1
                combined_entry["euclidean_distances"] = dst
                combined_data.append(combined_entry)

            step_data = {
                "iteration": iteration + 1,
                "count": k,
                "centroids": np.round(centroids, 2).tolist(),
                "clusters": (clusters).tolist(),
                "euclidean_distances": distances.tolist(),
                "combined_data": combined_data,
            }
            iteration_steps.append(step_data)

            # Update centroid
            new_centroids = np.array(
                [
                    (
                        X[clusters == i].mean(axis=0)
                        if np.any(clusters == i)
                        else centroids[i]
                    )
                    for i in range(k)
                ]
            )
            if np.all(
                new_centroids == centroids
            ):  # Jika tidak ada perubahan centroid, berhenti
                break
            centroids = new_centroids

        # Tambahkan hasil akhir ke DataFrame
        df["Kluster"] = clusters
        result.append(step_data)
        
        combined_data = result[0]["combined_data"]
        df = pd.DataFrame(combined_data)
        sorted_df = df.sort_values(by=["clusters"]).to_dict(orient="records")

        Kluster = [1,2,3,4]
        labels = ["Rendah","Unggul","Baik","Cukup Baik"]
        df["Euclidean Distance"] = [distances[i, clusters[i]] for i in range(len(X))]
        figure = generateKmeansFigur(X, (clusters).tolist(), centroids)
        figure_path = await figure
        # Menampilkan hasil ke console
        # print("\n=== Hasil K-Means Clustering ===")
        # for step in iteration_steps:
        #    print(f"\n--- Iterasi {step['iteration']} ---")
        #    print("Centroids:", step['centroids'])
        #    print("Clusters:", step['clusters'])
        #    print("Euclidean Distances:", step['euclidean_distances'])
        #    print("Euclidean Distances:", step['combined_data'])
        # Kirim data langkah-langkah ke template
        
        KlusterLabels = zip(Kluster, labels)
        return render(
            request, "page/kmeans_result.html", {"KlusterLabels":KlusterLabels,"result": sorted_df, "figure": figure_path}
        )

def label_cluster(cluster):
    if cluster == 0:
        return 'Rendah'
    elif cluster == 1:
        return 'Unggul'
    elif cluster == 2:
        return 'Baik'
    elif cluster == 3:
        return 'Cukup Baik'

async def generateKmeansFigur(X, labels, centroids):

    # Warna untuk tiap cluster
    colors = ["red", "green", "blue", "cyan", "purple", "yellow", "black", "orange", "teal", "pink", "indigo"]
    # Plot data points
    for cluster_id in np.unique(labels):
        plt.scatter(
            X[labels == cluster_id, 0],
            X[labels == cluster_id, 1],
            s=30,  # Ukuran marker
            color=colors[cluster_id],
            label=f"Kluster {cluster_id + 1}",
        )

    # Plot centroids
    plt.scatter(
        centroids[:, 0],
        centroids[:, 1],
        s=200,  # Ukuran marker centroid
        color="yellow",
        edgecolor="black",
        label="Centroid",
    )

    # Judul dan Label
    plt.title("Cluster Siswa Unggulan")
    plt.xlabel("Nilai Terbaik")
    plt.ylabel("Persentase Nilai")
    plt.legend()  # Menampilkan legenda
    plt.grid(True)

    new_id = max_id + 1
    figure_filename = f"kmeans_plot_{new_id}.png"
    # Make sure the directory exists
    os.makedirs(kmeans_dir, exist_ok=True)
    # Path to save the plot
    figure_path = os.path.join(kmeans_dir, figure_filename)

    print(figure_path)
    plt.savefig(figure_path, format="png")
    plt.close()

    figure_path = "kmeans/"+ figure_filename
    return figure_path


async def genarteElbowPlot(file, start_cluster, end_cluster, max_iter):

    context = {"form": ExcelUploadForm()}

    # Membuat DataFrame
    df = pd.read_excel(file)
    if df is None:
        context["error"] = "File Excel tidak valid atau kosong."
        return context
    # Pastikan dataset memiliki kolom yang sesuai
    required_columns = ["No", "X", "Y"]

    if not all(col in df.columns for col in required_columns):
        context["error"] = (
            f"Dataset harus memiliki kolom: {', '.join(required_columns)}"
        )
        return context
    X = df[["X", "Y"]].to_numpy()

    # Mengambil parameter dari form
    if start_cluster is None or end_cluster is None or max_iter is None:
        context["error"] = "Silahkan Isi Form dengan Benar"
        return context
    # Inisialisasi list untuk menyimpan nilai inertia
    inertia_values = []
    # Lakukan K-Means untuk berbagai jumlah cluster (start_cluster hingga end_cluster)
    for k in range(start_cluster, end_cluster):
        kmeans = KMeans(n_clusters=k, max_iter=max_iter, random_state=42)
        kmeans.fit(X)
        inertia_values.append(kmeans.inertia_)  # Menyimpan nilai inertia

    # Buat plot Elbow
    plt.figure()
    plt.plot(range(start_cluster, end_cluster), inertia_values, marker="o")
    plt.title("Elbow Method For Optimal k")
    plt.xlabel("Jumlah Cluster")
    plt.ylabel("WCSS")

    new_id = max_id + 1
    plot_filename = f"elbow_plot_{new_id}.png"
    # Make sure the directory exists
    os.makedirs(elbow_dir, exist_ok=True)

    # Path to save the plot
    plot_filepath = os.path.join(elbow_dir, plot_filename)

    print(plot_filepath)
    plt.savefig(plot_filepath, format="png")
    plt.close()
    
    # Kembalikan path relatif untuk digunakan di template
    plot_url = "elbow/"+ plot_filename
    return plot_url

# Function to convert image to base64
def image_to_base64(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode('utf-8')
    
async def generatePdf(table_data,name,elbow_file,figure_file,c_type,n_cluster,n_iter,s_elbow,e_elbow,n_elbow_iter):
     # Convert images to base64
    figure_base64 = image_to_base64(figure_file)
    elbow_base64 = image_to_base64(elbow_file)

    # Create an image tag using the base64 data
    figure_img_tag = f'data:image/png;base64,{figure_base64}'
    elbow_img_tag = f'data:image/png;base64,{elbow_base64}'
    new_id = max_id+1
    data_for_pdf = {
           "result": table_data,
           "init_method": c_type,
           "k": n_cluster,
           "n_iter": n_iter,
           "start_cluster": s_elbow,
           "end_cluster": e_elbow,
           "max_iter": n_elbow_iter,
           "elbow": figure_img_tag,
           "figure": elbow_img_tag
    }
    # Render the HTML content
    html_content = render_to_string('page/pdf_draw.html', data_for_pdf)
     # Create a BytesIO buffer to store the PDF
    buffer = BytesIO()
    # Create the PDF from the rendered HTML
    # Use xhtml2pdf to convert HTML to PDF
    pisa_status = pisa.CreatePDF(html_content, dest=buffer)

    # Check if PDF creation was successful
    if pisa_status.err:
        return HttpResponse('Error generating PDF', status=500)
     # Get the PDF data from the buffer
    pdf_data = buffer.getvalue()
    pdf_name = f"{new_id}_hasil_kmeans_{name}.pdf"
    
    # Make sure the directory exists
    os.makedirs(pdf_dir, exist_ok=True)
    
    # Save it inside MEDIA_ROOT
    file_path = os.path.join(pdf_dir, pdf_name)  # Saved under static/pdf/
    content_file = ContentFile(pdf_data)
    content_file.name = file_path
    with open(file_path, 'wb') as pdf_file:
        pdf_file.write(pdf_data)
        
    file_path = "pdf/"+ pdf_name
    
    return file_path

@sync_to_async
def save_clustering_data(name, pdf_file, c_type, n_cluster, n_iter, s_elbow, e_elbow, n_elbow_iter):
    
    new_id = max_id+1
    
    obj = m_data(
        id = new_id,
        data_name = name,
        dataset=pdf_file,  # Simpan konten file sebagai biner
        c_type=c_type,
        n_cluster=n_cluster,
        n_iter=n_iter,
        s_elbow=s_elbow,
        e_elbow=e_elbow,
        n_elbow_iter=n_elbow_iter,
    )
    obj.save()
    
    return obj



async def summary(request):

    context = {"form": ExcelUploadForm()}

    if request.method == "POST":
        try:
            file = request.FILES["file"]
        except Exception as e:
            context["error"] = "File Excel tidak valid atau kosong."
            return render(request, "page/elbow_draw.html", context)

        file = request.FILES["file"]

        # Membuat DataFrame
        df = pd.read_excel(file)
        if df is None:
            context["error"] = "File Excel tidak valid atau kosong."
            return render(request, "page/elbow_draw.html", context)
        # Pastikan dataset memiliki kolom yang sesuai
        required_columns = ["No", "X", "Y"]

        if not all(col in df.columns for col in required_columns):
            context["error"] = (
                f"Dataset harus memiliki kolom: {', '.join(required_columns)}"
            )
            return render(request, "page/elbow_draw", context)

        # Inisialisasi variabel untuk menyimpan langkah
        iteration_steps = []

        # Langkah 1: Inisialisasi centroid
        # Pilih kolom untuk clustering
        X = df[["X", "Y"]].to_numpy()

        try:
            name = request.POST.get("name") # Nama File
            k = int(request.POST.get("clusters", 10))  # Jumlah kluster
            n_iter = int(request.POST.get("iterations", 10))  # Jumlah Iterasi
            init_method = request.POST.get("init_method")  # Metode Insiasi
            start_cluster = int(request.POST.get("start_cluster", 10))  # Start Cluster
            end_cluster = int(request.POST.get("end_cluster", 10))  # End Cluster
            max_iter = int(request.POST.get("max_iter", 10))  # Jumlah Iterasi Elbow
        except Exception as e:
            context["error"] = "Silahkan Isi Form Inisiasi Cluster"
            return render(request, "page/elbow_draw.html", context)
        if k > 10:
            context["error"] = "Maximum Cluster 10"
            return render(request, "page/elbow_draw.html", context)
        # Manual K-Means untuk mencatat hasil per iterasi
        k = int(request.POST.get("clusters", 10))  # Jumlah kluster
        n_iter = int(request.POST.get("iterations", 10))  # Jumlah Iterasi
        init_method = request.POST.get("init_method")  # Metode Insiasi

        if init_method == "random":
            centroids = initialize_centroids_random(X, k)
        elif init_method == "average":
            centroids = initialize_centroids_average(X, k)
        else:
            context["error"] = "init_method must be 'random' or 'average'"
            return render(request, "page/elbow_draw.html", context)

        uploadedData = df.to_dict("records")

        # Inisialisasi variabel untuk menyimpan hasil
        result = []

        for iteration in range(n_iter):
            # Hitung jarak Euclidean
            distances = np.round(
                np.linalg.norm(X[:, None] - centroids[None, :, :], axis=2), 4
            )
            clusters = np.argmin(distances, axis=1)

            # Catat langkah per iterasi
            combined_data = []
            for data, cl, dst in zip(uploadedData, clusters, distances):
                combined_entry = data.copy()
                combined_entry["clusters"] = cl + 1
                combined_entry["euclidean_distances"] = dst
                combined_data.append(combined_entry)

            step_data = {
                "iteration": iteration + 1,
                "count": k,
                "centroids": np.round(centroids, 2).tolist(),
                "clusters": (clusters).tolist(),
                "euclidean_distances": distances.tolist(),
                "combined_data": combined_data,
            }
            iteration_steps.append(step_data)

            # Update centroid
            new_centroids = np.array(
                [
                    (
                        X[clusters == i].mean(axis=0)
                        if np.any(clusters == i)
                        else centroids[i]
                    )
                    for i in range(k)
                ]
            )
            if np.all(
                new_centroids == centroids
            ):  # Jika tidak ada perubahan centroid, berhenti
                break
            centroids = new_centroids

        # Tambahkan hasil akhir ke DataFrame
        result.append(step_data)
        figure = generateKmeansFigur(X, (clusters).tolist(), centroids)
        elbow = genarteElbowPlot(file, start_cluster, end_cluster, max_iter)
        figure_path = await figure
        elbow_path = await elbow
        
        def_figure_path = os.path.join(settings.BASE_DIR, "static", figure_path)
        def_elbow_path = os.path.join(settings.BASE_DIR, "static", elbow_path)
        await generatePdf(result, name,  def_figure_path,  def_elbow_path, init_method, k, n_iter, start_cluster, end_cluster, max_iter)
    
        new_id = max_id+1
        pdf_path = f"{new_id}_hasil_kmeans_{name}.pdf"

        await save_clustering_data(name, pdf_path, init_method, k, n_iter, start_cluster, end_cluster, max_iter)
        
        return render(
            request,
            "page/elbow_draw.html",
            {"result": result, "figure": figure_path, "elbow": elbow_path},
        )
        